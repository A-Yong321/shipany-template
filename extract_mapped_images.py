import openpyxl
import os
from PIL import Image as PILImage
import io

source_file = r'd:\shipany-template\图像资源.xlsx'
target_dir = r'd:\shipany-template\public\img_video\imgs'

# Ensure output directory exists
if not os.path.exists(target_dir):
    os.makedirs(target_dir)

print(f"Loading workbook: {source_file}")
try:
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    
    print(f"Found {len(ws._images)} images in worksheet.")

    # 1. Build a map of Row -> (Type, Scene) with fill-down logic
    # Find max row used
    max_row = ws.max_row
    row_map = {}
    current_b = None
    current_c = None
    
    print("Building metadata map...")
    for r in range(1, max_row + 1):
        b_val = ws.cell(row=r, column=2).value # Column B
        c_val = ws.cell(row=r, column=3).value # Column C
        
        if b_val:
            current_b = b_val
        if c_val:
            current_c = c_val
            
        row_map[r] = (current_b, current_c)

    count = 0
    for i, img in enumerate(ws._images):
        # Get anchor position
        # OneCellAnchor and TwoCellAnchor both usually have _from
        if not hasattr(img.anchor, '_from'):
            print(f"Img {i}: Unknown anchor")
            continue
            
        row = img.anchor._from.row + 1
        col = img.anchor._from.col + 1
        
        # We want Col D (4) -> _1 and Col E (5) -> _2
        # Debug: check rows/cols
        # print(f"Img {i} at {row},{col}")

        if col not in [4, 5]:
            continue
            
        # Lookup metadata
        b_val, c_val = row_map.get(row, (None, None))
        
        if not b_val or not c_val:
            print(f"Skipping image at Row {row}, Col {col}: Metadata missing. (Type={b_val}, Scene={c_val})")
            continue
            
        # Clean filenames
        def clean_name(s):
            return str(s).strip().replace(' ', '-').replace('/', '-').replace('\\', '-')
            
        b_clean = clean_name(b_val)
        c_clean = clean_name(c_val)
        
        suffix = None
        if col == 4: # Column D - 原图
            suffix = "1"
        elif col == 5: # Column E - 效果图
            suffix = "2"
        
        filename = f"{b_clean}_{c_clean}_{suffix}.png"
        save_path = os.path.join(target_dir, filename)
        
        # Save logic
        try:
            saved = False
            # Method 1: img.ref (path)
            if hasattr(img, 'ref'):
                 try:
                     pil_img = PILImage.open(img.ref)
                     pil_img.save(save_path)
                     saved = True
                 except:
                     pass
            
            # Method 2: img.image (embedded)
            if not saved and hasattr(img, 'image'):
                 try:
                     image_bytes = img.image._data()
                     pil_img = PILImage.open(io.BytesIO(image_bytes))
                     pil_img.save(save_path)
                     saved = True
                 except:
                     pass

            # Method 3: img._data (direct)
            if not saved:
                try:
                    image_bytes = img._data()
                    pil_img = PILImage.open(io.BytesIO(image_bytes))
                    pil_img.save(save_path)
                    saved = True
                except:
                    pass

            if saved:
                print(f"Saved {filename}")
                count += 1
            else:
                print(f"Failed to extract content for {filename}")

        except Exception as save_err:
            print(f"Error saving {filename}: {save_err}")

    print(f"Extraction finished. Saved {count} images.")

except Exception as e:
    print(f"Process failed: {e}")
