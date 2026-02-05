
import openpyxl
from openpyxl.drawing.image import Image
import os

def export_images(excel_path, output_dir):
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print(f"Active sheet: {ws.title}")

    # Map images to their anchor cells (row, col)
    # openpyxl images are stored in ws._images
    images_in_cells = {}
    
    print(f"Found {len(ws._images)} images in the worksheet.")

    for img in ws._images:
        # Get anchor position
        # Note: Anchor implementation can vary (OneCell vs TwoCell)
        # We assume top-left corner is what matters
        row = img.anchor._from.row + 1 # 0-indexed to 1-indexed
        col = img.anchor._from.col + 1 # 0-indexed to 1-indexed
        
        if row not in images_in_cells:
            images_in_cells[row] = {}
        images_in_cells[row][col] = img

    # Iterate through rows and extract images based on requirements
    # B=2, C=3, D=4, F=6
    processed_count = 0
    
    # Iterate through rows that have images or data
    # We can perform a scan based on the max row in the sheet
    for row in range(1, ws.max_row + 1):
        # Get values for naming: Column B (2) and C (3)
        b_val = ws.cell(row=row, column=2).value
        c_val = ws.cell(row=row, column=3).value

        # Only proceed if we have valid naming components
        if b_val and c_val:
            b_str = str(b_val).strip()
            c_str = str(c_val).strip()
            
            # Check for image in Column D (4)
            if row in images_in_cells and 4 in images_in_cells[row]:
                img = images_in_cells[row][4]
                filename = f"{b_str}_{c_str}_1.png"
                save_path = os.path.join(output_dir, filename)
                # Save image
                # img.ref is a BytesIO object or similar, verify specific openpyxl version behavior
                # Usually img.image is the PIL Image object (if extracted correctly) or we can save bytes
                
                try:
                    # Depending on openpyxl version/image type (PIL image vs others)
                    
                    # If it's a PIL based image
                    if hasattr(img, 'image'): # Does it hold a PIL image directly?
                        img.image.save(save_path)
                    else:
                        # Fallback or standard way: img.ref might be a file-like object or content
                         with open(save_path, "wb") as f:
                            f.write(img._data())
                    
                    print(f"Saved {filename}")
                    processed_count += 1
                except Exception as e:
                     # Attempt to use the pillow image object if available
                    try:
                        # Attempt to access underlying data if previous methods failed
                        if hasattr(img, 'ref') and hasattr(img.ref, 'getvalue'):
                             with open(save_path, "wb") as f:
                                 f.write(img.ref.getvalue())
                             print(f"Saved {filename} (fallback)")
                             processed_count += 1
                        else:
                             print(f"Error saving {filename}: {e}")
                    except Exception as e2:
                        print(f"Error saving {filename}: {e}. Fallback failed: {e2}")

            # Check for image in Column F (6)
            if row in images_in_cells and 6 in images_in_cells[row]:
                img = images_in_cells[row][6]
                filename = f"{b_str}_{c_str}_2.png"
                save_path = os.path.join(output_dir, filename)
                
                try:
                     # Simpler save approach for modern openpyxl
                    if hasattr(img, 'ref'):
                         # img.ref is usually a BytesIO
                         with open(save_path, "wb") as f:
                             f.write(img.ref.getvalue())
                    else:
                        # Backup: access _data
                        with open(save_path, "wb") as f:
                             f.write(img._data())
                    print(f"Saved {filename}")
                    processed_count += 1
                except Exception as e:
                    print(f"Error saving {filename}: {e}")

    print(f"Export complete. Processed {processed_count} images.")

if __name__ == "__main__":
    # Define paths
    # Using absolute paths or relative to script execution
    # Script is in scripts/, so root is one up.. or we just use absolute paths provided in prompt context
    
    # User said: @[图像资源.xlsx] -> d:\shipany-template\图像资源.xlsx
    # Dest: @[public/img_video/imgs] -> d:\shipany-template\public\img_video\imgs
    
    excel_file = r"d:\shipany-template\图像资源.xlsx"
    dest_dir = r"d:\shipany-template\public\img_video\imgs"
    
    export_images(excel_file, dest_dir)
