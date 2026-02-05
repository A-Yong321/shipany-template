"""
检查Excel中所有图片的位置
"""
import openpyxl

source_file = r'd:\shipany-template\图像资源.xlsx'

try:
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    
    print(f"工作表名称: {ws.title}")
    print(f"找到 {len(ws._images)} 个图片\n")
    
    # 统计每列的图片数量
    col_count = {}
    
    for i, img in enumerate(ws._images):
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1
            col_letter = openpyxl.utils.get_column_letter(col)
            
            if col not in col_count:
                col_count[col] = 0
            col_count[col] += 1
            
            # 打印前10个图片的位置
            if i < 10:
                # 获取该行的B列和C列值
                b_val = ws.cell(row=row, column=2).value
                c_val = ws.cell(row=row, column=3).value
                print(f"图片 {i+1}: 行{row}, 列{col}({col_letter}) - B={b_val}, C={c_val}")
    
    print(f"\n每列的图片数量统计:")
    for col in sorted(col_count.keys()):
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  列 {col} ({col_letter}): {col_count[col]} 个图片")
    
    print(f"\n总计: {sum(col_count.values())} 个图片")

except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()
