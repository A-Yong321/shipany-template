import openpyxl
import os

source_file = r'd:\1PhotoAI\资源.xlsx'

try:
    wb = openpyxl.load_workbook(source_file, data_only=True)
    sheet = wb.active
    
    print(f"Sheet Name: {sheet.title}")
    
    # Print headers (Row 1)
    headers = []
    for col in range(1, 10):
        cell_val = sheet.cell(row=1, column=col).value
        headers.append(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): {cell_val}")
    print("Headers:", headers)
    
    # Print first 5 data rows
    print("\nFirst 5 data rows:")
    for row in range(2, 7):
        vals = []
        for col in range(1, 10):
            vals.append(f"{openpyxl.utils.get_column_letter(col)}={sheet.cell(row=row, column=col).value}")
        print(f"Row {row}: {', '.join(vals)}")
        
    # Check for images
    print(f"\nFound {len(sheet._images)} images.")
    if len(sheet._images) > 0:
        img = sheet._images[0]
        print(f"Sample Image Anchor: {img.anchor}")

except Exception as e:
    print(f"Error: {e}")
